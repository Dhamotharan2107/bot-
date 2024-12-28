# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# import time
# import os
# from openpyxl import load_workbook
# from tkinter import Tk
# from tkinter.filedialog import askopenfilename, askdirectory

# def send_images_to_whatsapp():
#     # File and folder selection
#     Tk().withdraw()
#     excel_path = askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
#     if not excel_path:
#         print("No Excel file selected. Exiting.")
#         return

#     root_folder = askdirectory(title="Select Root Folder for UQID Folders")
#     if not root_folder:
#         print("No folder selected. Exiting.")
#         return

#     # ChromeDriver setup
#     options = Options()
#     options.add_argument("profile-directory=Profile 1")
#     options.add_argument(r"C:\Users\THAAGAM\Desktop\WSC\chrome-data")  # Use raw string to avoid unicode error
#     options.add_experimental_option("excludeSwitches", ["enable-automation"])
#     options.add_experimental_option('useAutomationExtension', False)
    
#     # If you want to make sure that a fresh login is done every time, disable the use of the previous session:
#     # options.add_argument("--incognito")  # Use incognito mode to avoid using old sessions

#     # Start the browser with the given options
#     driver = webdriver.Chrome(options=options)
#     driver.get("http://web.whatsapp.com")

#     # Wait for the user to scan the QR code for login (always re-login on each run)
#     print("Please scan the QR code to log in to WhatsApp Web.")
#     input("Press Enter after scanning the QR code...")

#     # Load Excel
#     wb = load_workbook(excel_path)
#     sheet = wb.active

#     for row in range(2, sheet.max_row + 1):  # Skip headers
#         uqid = sheet.cell(row, 1).value
#         phone_number = sheet.cell(row, 2).value
#         status = sheet.cell(row, 3).value

#         # Skip if the status is "Success"
#         if status == "Success":
#             print(f"Skipping {phone_number} because the status is already Success.")
#             continue

#         # Skip rows with missing phone number or UQID
#         if not uqid or not phone_number:
#             print(f"Skipping row {row} due to missing UQID or phone number.")
#             sheet.cell(row, 3).value = "Missing UQID or Phone Number"
#             sheet.cell(row, 4).value = 0  # Set the pictures sent count to 0
#             continue

#         folder_path = os.path.join(root_folder, str(uqid))
#         if not os.path.exists(folder_path):
#             print(f"Folder not found for UQID: {uqid}")
#             sheet.cell(row, 3).value = "Folder Not Found"
#             sheet.cell(row, 4).value = 0  # Set the pictures sent count to 0
#             continue

#         # Check if there are any valid image files in the folder
#         image_files = [os.path.join(folder_path, f) for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f)) and f.lower().endswith(('png', 'jpg', 'jpeg'))]
#         if not image_files:
#             print(f"No valid images found in folder: {folder_path}")
#             sheet.cell(row, 3).value = "No Images Found"
#             sheet.cell(row, 4).value = 0  # Set the pictures sent count to 0
#             continue

#         try:
#             # Use direct WhatsApp URL to open the chat with phone number
#             webpage = f"https://web.whatsapp.com/send?phone={phone_number}"
#             driver.get(webpage)
#             time.sleep(5)  # Wait for the chat to open

#             # Check if the phone number is not registered on WhatsApp
#             try:
#                 # If the phone number is not registered, a prompt with a message will appear
#                 WebDriverWait(driver, 10).until(
#                     EC.presence_of_element_located((By.XPATH, "//div[@class='_1XilN']"))  # This is the "Not on WhatsApp" message.
#                 )
#                 print(f"{phone_number} is not on WhatsApp.")
#                 sheet.cell(row, 3).value = "Incorrect Number"
#                 sheet.cell(row, 4).value = 0  # Set the pictures sent count to 0
#                 continue
#             except Exception:
#                 pass  # If the "Not on WhatsApp" message is not found, proceed with sending images

#             # Print to indicate the attach process
#             print("Attach")

#             # Wait for the 'Attach' button to appear and click it
#             WebDriverWait(driver, 50).until(
#                 EC.presence_of_element_located((By.XPATH, "//*[@id='main']/footer/div[1]/div/span/div/div[1]/div[2]/button/span"))
#             )
#             attach_button = driver.find_element(By.XPATH, "//*[@id='main']/footer/div[1]/div/span/div/div[1]/div[2]/button/span")
#             attach_button.click()
#             time.sleep(2)

#             # Wait for the file input element for photos/videos to appear
#             WebDriverWait(driver, 50).until(
#                 EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']"))
#             )

#             # Send all images in one go
#             photo_input = driver.find_element(By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']")
#             # Join all image paths with newline to simulate sending all files at once
#             photo_input.send_keys("\n".join(image_files))
#             time.sleep(2)

#             # Wait for the 'Send' button to become clickable after uploading the photos
#             WebDriverWait(driver, 130).until(
#                 EC.element_to_be_clickable((By.XPATH, "//span[@data-icon='send']"))
#             )
#             send_button = driver.find_element(By.XPATH, "//span[@data-icon='send']")
#             send_button.click()
#             time.sleep(10)

#             # Update the "Pictures Sent" count in the Excel sheet
#             sheet.cell(row, 4).value = len(image_files)  # Column 4 is used to store the count
#             print(f"Images sent successfully to {phone_number}, Count: {len(image_files)}")
#             sheet.cell(row, 3).value = "Success"

#         except Exception as e:
#             print(f"Failed to send images to {phone_number}: {e}")
#             sheet.cell(row, 3).value = f"Failure: {e}"
#             sheet.cell(row, 4).value = 0  # Set the pictures sent count to 0 in case of failure

#     wb.save(excel_path)
#     driver.quit()
#     print("Process completed. Check the Excel file for status updates.")

# send_images_to_whatsapp()


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory

def send_images_to_whatsapp():
    # File and folder selection
    Tk().withdraw()
    excel_path = askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        print("No Excel file selected. Exiting.")
        return

    root_folder = askdirectory(title="Select Root Folder for UQID Folders")
    if not root_folder:
        print("No folder selected. Exiting.")
        return

    # Extract UQIDs from the folder
    uqid_folders = [name for name in os.listdir(root_folder) if os.path.isdir(os.path.join(root_folder, name))]

    # Load Excel
    wb = load_workbook(excel_path)
    sheet = wb.active

    # ChromeDriver setup
    options = Options()
    options.add_argument("profile-directory=Profile 1")
    options.add_argument(r"C:\Users\THAAGAM\Desktop\WSC\chrome-data")  # Use raw string to avoid unicode error
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    # Start the browser with the given options
    driver = webdriver.Chrome(options=options)
    driver.get("http://web.whatsapp.com")

    # Wait for the user to scan the QR code
    print("Please scan the QR code to log in to WhatsApp Web.")
    input("Press Enter after scanning the QR code...")

    for folder_uqid in uqid_folders:
        # Check if UQID exists in the Excel sheet
        found = False
        for row in range(2, sheet.max_row + 1):
            uqid = str(sheet.cell(row, 1).value).strip()
            phone_number = sheet.cell(row, 2).value
            status = sheet.cell(row, 3).value

            if uqid == folder_uqid:
                found = True
                # Skip if already successful
                if status == "Success":
                    print(f"Skipping {phone_number} because the status is already Success.")
                    continue

                folder_path = os.path.join(root_folder, folder_uqid)
                # Check for valid image files
                image_files = [
                    os.path.join(folder_path, f) for f in os.listdir(folder_path)
                    if os.path.isfile(os.path.join(folder_path, f)) and f.lower().endswith(('png', 'jpg', 'jpeg'))
                ]
                if not image_files:
                    print(f"No valid images found in folder: {folder_path}")
                    sheet.cell(row, 3).value = "No Images Found"
                    sheet.cell(row, 4).value = 0
                    continue

                try:
                    # Open WhatsApp chat
                    webpage = f"https://web.whatsapp.com/send?phone={phone_number}"
                    driver.get(webpage)
                    time.sleep(5)

                    # Check if the number is not on WhatsApp
                    try:
                        WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, "//div[@class='_1XilN']"))
                        )
                        print(f"{phone_number} is not on WhatsApp.")
                        sheet.cell(row, 3).value = "Incorrect Number"
                        sheet.cell(row, 4).value = 0
                        continue
                    except Exception:
                        pass

                    # Attach and send images
                    WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, "//*[@id='main']/footer/div[1]/div/span/div/div[1]/div[2]/button/span"))
                    ).click()

                    photo_input = WebDriverWait(driver, 50).until(
                        EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']"))
                    )
                    photo_input.send_keys("\n".join(image_files))
                    time.sleep(2)

                    WebDriverWait(driver, 130).until(
                        EC.element_to_be_clickable((By.XPATH, "//span[@data-icon='send']"))
                    ).click()
                    time.sleep(5)

                    # Update Excel
                    sheet.cell(row, 3).value = "Success"
                    sheet.cell(row, 4).value = len(image_files)
                    print(f"Images sent successfully to {phone_number}. Count: {len(image_files)}")
                except Exception as e:
                    print(f"Failed to send images to {phone_number}: {e}")
                    sheet.cell(row, 3).value = f"Failure: {e}"
                    sheet.cell(row, 4).value = 0

        if not found:
            print(f"UQID {folder_uqid} not found in Excel. Skipping.")
    
    # Save updates to Excel
    wb.save(excel_path)
    driver.quit()
    print("Process completed. Check the Excel file for status updates.")

send_images_to_whatsapp()
