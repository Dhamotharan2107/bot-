from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename, askdirectory

def send_images_to_whatsapp():
    # File and folder selection
    Tk().withdraw()
    excel_path = askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        print("No Excel file selected. Exiting.")
        return

    root_folder = askdirectory(title="Select Root Folder for UQID Folders")
    if not root_folder:
        print("No folder selected. Exiting.")
        return

    # Extract UQIDs from the folder
    uqid_folders = [name for name in os.listdir(root_folder) if os.path.isdir(os.path.join(root_folder, name))]

    # Load Excel
    wb = load_workbook(excel_path)
    sheet = wb.active

    # ChromeDriver setup
    options = Options()
    profile_path = r"C:\Users\THAAGAM\Desktop\WSC\chrome-data"  # Folder to store the Chrome profile data
    options.add_argument(f"user-data-dir={profile_path}")  # Set the user data directory
    options.add_argument("profile-directory=Profile 1")  # Use a specific profile
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])  # Disable automation flag
    options.add_experimental_option('useAutomationExtension', False)
    
    # Start the browser with the given options
    driver = webdriver.Chrome(options=options)
    driver.get("http://web.whatsapp.com")

    # Wait for the user to scan the QR code
    print("Please scan the QR code to log in to WhatsApp Web.")
    input("Press Enter after scanning the QR code...")

    for folder_uqid in uqid_folders:
        # Check if UQID exists in the Excel sheet
        found = False
        for row in range(2, sheet.max_row + 1):
            uqid = str(sheet.cell(row, 1).value).strip()
            phone_number = sheet.cell(row, 2).value
            status = sheet.cell(row, 3).value

            if uqid == folder_uqid:
                found = True
                # Skip if already successful
                if status == "Success":
                    print(f"Skipping {phone_number} because the status is already Success.")
                    continue

                folder_path = os.path.join(root_folder, folder_uqid)
                # Check for valid image files
                image_files = [
                    os.path.join(folder_path, f) for f in os.listdir(folder_path)
                    if os.path.isfile(os.path.join(folder_path, f)) and f.lower().endswith(('png', 'jpg', 'jpeg'))
                ]
                if not image_files:
                    print(f"No valid images found in folder: {folder_path}")
                    sheet.cell(row, 3).value = "No Images Found"
                    sheet.cell(row, 4).value = 0
                    continue

                try:
                    # Open WhatsApp chat
                    webpage = f"https://web.whatsapp.com/send?phone={phone_number}"
                    driver.get(webpage)

                    # Check if the number is not on WhatsApp
                    try:
                        WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//div[@class='_1XilN']"))
                        )
                        print(f"{phone_number} is not on WhatsApp.")
                        sheet.cell(row, 3).value = "Incorrect Number"
                        sheet.cell(row, 4).value = 0
                        continue
                    except Exception:
                        pass

                    # Attach and send images
                    WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, "//*[@id='main']/footer/div[1]/div/span/div/div[1]/div[2]/button/span"))
                    ).click()

                    photo_input = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']"))
                    )
                    photo_input.send_keys("\n".join(image_files))  # Send all images at once
                    time.sleep(1)

                    WebDriverWait(driver, 20).until(
                        EC.element_to_be_clickable((By.XPATH, "//span[@data-icon='send']"))
                    ).click()
                    time.sleep(5)  # Shortened wait after sending

                    # Update Excel
                    sheet.cell(row, 3).value = "Success"
                    sheet.cell(row, 4).value = len(image_files)
                    print(f"Images sent successfully to {phone_number}. Count: {len(image_files)}")
                except Exception as e:
                    print(f"Failed to send images to {phone_number}: {e}")
                    sheet.cell(row, 3).value = f"Failure: {e}"
                    sheet.cell(row, 4).value = 0

        if not found:
            print(f"UQID {folder_uqid} not found in Excel. Skipping.")
    
    # Save updates to Excel
    wb.save(excel_path)
    driver.quit()
    print("Process completed. Check the Excel file for status updates.")

send_images_to_whatsapp()
