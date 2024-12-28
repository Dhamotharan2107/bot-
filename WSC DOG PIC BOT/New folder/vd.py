from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename

def send_video_to_whatsapp():
    # File and folder selection
    Tk().withdraw()
    excel_path = askopenfilename(title="Select Excel File", filetypes=[("Excel Files", "*.xlsx")])
    if not excel_path:
        print("No Excel file selected. Exiting.")
        return

    video_path = askopenfilename(title="Select Video File", filetypes=[("Video Files", "*.mp4;*.avi;*.mov")])
    if not video_path:
        print("No video file selected. Exiting.")
        return

    # Load Excel
    wb = load_workbook(excel_path)
    sheet = wb.active

    # ChromeDriver setup
    options = Options()
    profile_path = r"C:\Users\THAAGAM\Desktop\WSC\chrome-data"  # Folder to store the Chrome profile data
    options.add_argument(f"user-data-dir={profile_path}")  # Set the user data directory
    options.add_argument("profile-directory=Profile 1")  # Use a specific profile
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])  # Disable automation flag
    options.add_experimental_option('useAutomationExtension', False)
    
    # Start the browser with the given options
    driver = webdriver.Chrome(options=options)
    driver.get("http://web.whatsapp.com")

    # Wait for the user to scan the QR code
    print("Please scan the QR code to log in to WhatsApp Web.")
    input("Press Enter after scanning the QR code...")

    # Iterate through phone numbers in the Excel file
    for row in range(2, sheet.max_row + 1):
        phone_number = str(sheet.cell(row, 1).value).strip()
        status = sheet.cell(row, 2).value  # Status column

        # Skip if already successful
        if status == "Success":
            print(f"Skipping {phone_number} because the status is already Success.")
            continue

        try:
            # Open WhatsApp chat
            webpage = f"https://web.whatsapp.com/send?phone={phone_number}"
            driver.get(webpage)

            # Check if the number is not on WhatsApp
            try:
                WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, "//div[@class='_1XilN']"))
                )
                print(f"{phone_number} is not on WhatsApp.")
                sheet.cell(row, 2).value = "Incorrect Number"
                continue
            except Exception:
                pass

            # Attach and send video as a document
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='main']/footer/div[1]/div/span/div/div[1]/div[2]/button/span"))
            ).click()

            # Select Document Option
            WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/span[5]/div/ul/div/div/div[1]/li"))
            ).click()

            # Upload the video file
            document_input = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))
            )
            document_input.send_keys(video_path)  # Send the selected video file
            time.sleep(2)

            # Click the send button
            WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='app']/div/div[3]/div[2]/div[2]/span/div/div/div/div[2]/div/div[2]/div[2]/div/div"))
            ).click()

            time.sleep(15)  # Wait for the video to be sent

            # Update Excel with status
            sheet.cell(row, 2).value = "Success"
            print(f"Video sent successfully to {phone_number}.")

        except Exception as e:
            print(f"Failed to send video to {phone_number}: {e}")
            sheet.cell(row, 2).value = f"Failure: {e}"

    # Save updates to Excel
    wb.save(excel_path)
    driver.quit()
    print("Process completed. Check the Excel file for status updates.")

send_video_to_whatsapp()
